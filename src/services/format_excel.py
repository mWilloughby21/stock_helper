# format_excel.py

import pandas_market_calendars as mcal
from openpyxl.styles import Alignment
from datetime import datetime


class SetupFormatter:
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def center(self):
        for row in self.ws.iter_rows(min_row=1, max_row=260, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def get_dates(self, year=None):
        year = year or datetime.now().year
        nyse = mcal.get_calendar('NYSE')
        schedule = nyse.schedule(start_date=f'{year}-01-01', end_date=f'{year}-12-31')
        
        return list(zip(schedule.index.date, schedule.index.day_name()))
    
    def get_year(self):
        self.ws["B3"] = datetime.now().year
    
    def specific_cells(self):
        self.ws.cell(row=2, column=2, value="Year")
        self.ws.cell(row=2, column=1, value="First Trade Date")
        self.ws.cell(row=3, column=1, value=self.get_dates()[0][0])
        self.ws.cell(row=5, column=1, value="Date")
        self.ws.cell(row=5, column=2, value="Day")
    
    # Formatting 
    def static_format(self):
        self._set_column_widths()
        self._freeze_panes()
    
    def _set_column_widths(self):
        widths = {
            "A": 14,  # Market Dates
            "B": 8,  # Day of Week
        }
        
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width
    
    def _freeze_panes(self):
        self.ws.freeze_panes = "C1"