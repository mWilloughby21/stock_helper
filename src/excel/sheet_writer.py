# sheet_writer.py

START_COL = "C"

from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

class SheetWriter:
    def __init__(self, ws: Worksheet):
        self.ws = ws
    
    def update_stocks(self, cols, stock):
        pass
    
    def write_info(self, cols, stock):
        pass
    
    def get_block(self):
        col = column_index_from_string(START_COL)
        while self.ws.cell(row=1, column=col).value is not None: # 7
            col += 5
        return [num for num in range(col, col + 5)]
    
    def format_info(self, cols):
        self._set_col_widths(cols)
        self._center(cols)
        thin = Side(style="thin")
        hair = Side(style="hair")
        
        self.ws.cell(row=1, column=cols[0]).value = "Cost Basis"
        self.ws.cell(row=1, column=cols[1]).border = Border(left=thin, right=thin, bottom=thin, top=thin)
        self.ws.cell(row=1, column=cols[3]).value = "Avg $/Sh"
        self.ws.cell(row=1, column=cols[4]).border = Border(left=thin, right=thin, bottom=thin, top=thin)
        
        self.ws.cell(row=3, column=cols[0]).value = "Shares"
        self.ws.cell(row=3, column=cols[1]).border = Border(left=thin, right=thin, bottom=thin, top=thin)
        self.ws.cell(row=3, column=cols[3]).value = "Port %"
        self.ws.cell(row=3, column=cols[4]).border = Border(left=thin, right=thin, bottom=thin, top=thin)
        
        for col in cols:
            self.ws.cell(row=5, column=col).border = Border(top=thin)
            self.ws.cell(row=6, column=col).border = Border(top=thin)
            self.ws.cell(row=7, column=col).border = Border(left=thin, right=thin, bottom=thin, top=thin)
        
        self.ws.cell(row=2, column=cols[4]).border = Border(right=thin)
        self.ws.cell(row=4, column=cols[4]).border = Border(right=thin)
        self.ws.cell(row=5, column=cols[4]).border = Border(right=thin, top=thin)
        self.ws.cell(row=6, column=cols[4]).border = Border(right=thin, top=thin)
        
        self.ws.cell(row=7, column=cols[1]).value = "% Change"
        self.ws.cell(row=7, column=cols[3]).value = "Avg Cost ∆"
        self.ws.cell(row=7, column=cols[4]).value = "YTD ∆"
        
        for row in range(8, 260):
            for col in cols[:4]:
                self.ws.cell(row=row, column=col).border = Border(right=hair)
            self.ws.cell(row=row, column=cols[4]).border = Border(right=thin)
    
    def _center(self, cols):
        for row in self.ws.iter_rows(min_row=1, max_row=260, min_col=cols[0], max_col=cols[4]):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _set_col_widths(self, cols):
        for col in cols:
            self.ws.column_dimensions[get_column_letter(col)].width = 10