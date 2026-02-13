# workbook_manager.py

from openpyxl import Workbook, load_workbook
from pathlib import Path

class WorkbookManager:
    def __init__(self, path: str):
        self.path = Path(path)
        self.wb: Workbook | None = None
    
    def load_or_create(self):
        if self.path.exists():
            self.wb = load_workbook(self.path)
        else:
            self.wb = Workbook()
        return self.wb
    
    def save(self):
        self.wb.save(self.path)